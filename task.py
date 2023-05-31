import urllib.request
import re
from openpyxl import Workbook
from RPA.Browser.Selenium import Selenium
import json


class WebScraper:
    """
    Class for performing web scraping actions using Selenium.
    """

    def __init__(self):
        self.browser = Selenium()

    def open_browser(self, url):
        """
        Opens a browser and navigates to the specified URL.
        """
        self.browser.open_available_browser(url)

    def click_search_button(self, button_search):
        """
        Clicks on the search button to enable the search field.
        """
        self.browser.click_button(button_search)

    def input_text(self, text_area, text):
        """
        Inputs text into the specified text area.
        """
        self.browser.input_text(text_area, text)

    def click_go_button(self, button_go):
        """
        Clicks on the go button to start the search.
        """
        self.browser.click_button(button_go)

    def find_elements(self, xpath):
        """
        Finds web elements based on the specified XPath.
        """
        return self.browser.find_elements(xpath)

    def wait_until_element_present(self, xpath):
        """
        Waits until the specified element is present on the page.
        """
        self.browser.wait_until_page_contains_element(xpath)

    def get_text(self, element):
        """
        Retrieves the text of a web element.
        """
        return self.browser.get_text(element)

    def close_browser(self):
        """
        Closes the browser.
        """
        self.browser.close_all_browsers()


class ExcelWriter:
    """
    Class for writing data to an Excel file.
    """

    def __init__(self, headers):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.headers = headers

    def write_headers(self):
        """
        Writes the headers to the Excel file.
        """
        for col, head in enumerate(self.headers, start=1):
            cell = self.ws.cell(row=1, column=col)
            cell.value = head

    def write_title(self, row, title):
        """
        Writes the title to the specified row in the Excel file.
        """
        cell = self.ws.cell(row=row, column=1)
        cell.value = title

    def write_date(self, row, date):
        """
        Writes the date to the specified row in the Excel file.
        """
        cell = self.ws.cell(row=row, column=2)
        cell.value = date

    def write_description(self, row, description):
        """
        Writes the description to the specified row in the Excel file.
        """
        cell = self.ws.cell(row=row, column=3)
        cell.value = description

    def write_image_name(self, row, image_name):
        """
        Writes the image name to the specified row in the Excel file.
        """
        cell = self.ws.cell(row=row, column=4)
        cell.value = image_name

    def save_excel(self, file_path):
        """
        Saves the Excel file.
        """
        self.wb.save(file_path)


class NewsItem:
    """
    Class for storing the data of a news item.
    """

    def __init__(self, title, date, description, image_name):
        self.title = title
        self.date = date
        self.description = description
        self.image_name = image_name


def load_config_from_json(file_path):
    """
    Loads and returns the configuration data from a JSON file.
    """
    with open(file_path, 'r') as json_file:
        data = json.load(json_file)
    return data


def count_phrase(phrase, text):
    """
    Counts the number of times the keyword appears in the text.
    """
    count = text.lower().count(phrase.lower())
    return count


def find_money_formats(text):
    """
    Finds the correct $ format in the text.
    """
    pattern = r'\$[\d,]+(\.\d+)?|\d+ dollars|\d+ USD'
    matches = re.findall(pattern, text)
    return matches


def assign_boolean_value(bool_value, cell):
    """
    Assigns True or False based on whether the keyword appears.
    """
    if bool_value:
        if cell.value is None or cell.value is False:
            cell.value = True
    else:
        cell.value = False


def go_and_search(web_scraper, url, button_search, text_area, search_phrase, button_go):
    """
    Performs the search on the website.
    """
    web_scraper.open_browser(url)
    web_scraper.click_search_button(button_search)
    web_scraper.input_text(text_area, search_phrase)
    web_scraper.click_go_button(button_go)


def filter_category(web_scraper, news_category):
    """
    Applies the specified category filter.
    """
    checkboxes = web_scraper.find_elements("css:.css-1qtb2wd label.css-1a8ayg6")

    for checkbox in checkboxes:
        checkbox_text = web_scraper.get_text(checkbox)
        if news_category.lower() in checkbox_text.lower():
            web_scraper.click(checkbox)
            break


def extract_elements_titles(web_scraper, titles_xpath):
    """
    Extracts the titles of the news items.
    """
    web_scraper.wait_until_element_present(titles_xpath)
    titles = web_scraper.find_elements(titles_xpath)
    return titles


def extract_elements_dates(web_scraper, dates_xpath):
    """
    Extracts the dates of the news items.
    """
    web_scraper.wait_until_element_present(dates_xpath)
    dates = web_scraper.find_elements(dates_xpath)
    return dates


def extract_elements_descriptions(web_scraper, descriptions_xpath):
    """
    Extracts the descriptions of the news items.
    """
    web_scraper.wait_until_element_present(descriptions_xpath)
    descriptions = web_scraper.find_elements(descriptions_xpath)
    return descriptions


def extract_elements_images(web_scraper, images_xpath):
    """
    Extracts the images of the news items.
    """
    web_scraper.wait_until_element_present(images_xpath)
    images = web_scraper.find_elements(images_xpath)
    return images


def open_excel(writer, news_items, search_phrase):
    """
    Writes the news item data to an Excel file.
    """
    writer.write_headers()

    for row, news_item in enumerate(news_items, start=2):
        writer.write_title(row, news_item.title)
        count = count_phrase(search_phrase, news_item.title)
        writer.write_description(row, news_item.description)
        count += count_phrase(search_phrase, news_item.description)
        writer.write_date(row, news_item.date)
        writer.write_image_name(row, news_item.image_name)
        matches = find_money_formats(news_item.title)
        matches += find_money_formats(news_item.description)
        assign_boolean_value(bool(matches), writer.ws.cell(row=row, column=6))

        count_cell = writer.ws.cell(row=row, column=5)
        count_cell.value = count_cell.value + count if count_cell.value else count


def main():
    config = load_config_from_json('work_items/config.json')

    search_phrase = config.get('search_phrase', 'time')
    news_category = config.get('news_category', '')
    num_months = config.get('num_months', 1)
    url = config.get('url', '')
    titles_xpath = config.get('titles_xpath', '')
    dates_xpath = config.get('dates_xpath', '')
    descriptions_xpath = config.get('descriptions_xpath', '')
    images_xpath = config.get('images_xpath', '')
    button_search = config.get('button_search', '')
    text_area = config.get('text_area', '')
    button_go = config.get('button_go', '')
    file_path = config.get('file_path', '')

    web_scraper = WebScraper()
    web_scraper.open_browser(url)
    go_and_search(web_scraper, url, button_search, text_area, search_phrase, button_go)
    filter_category(web_scraper, news_category)

    titles = extract_elements_titles(web_scraper, titles_xpath)
    dates = extract_elements_dates(web_scraper, dates_xpath)
    descriptions = extract_elements_descriptions(web_scraper, descriptions_xpath)
    images = extract_elements_images(web_scraper, images_xpath)

    news_items = []
    for title, date, description, image in zip(titles, dates, descriptions, images):
        title_text = web_scraper.get_text(title)
        date_text = web_scraper.get_text(date)
        description_text = web_scraper.get_text(description)
        image_name = image.get_attribute("src").split("/")[-1]
        news_item = NewsItem(title_text, date_text, description_text, image_name)
        news_items.append(news_item)

    web_scraper.close_browser()

    writer = ExcelWriter(['Title', 'Date', 'Description', 'Image Name', 'Count', 'Has Money Format'])
    open_excel(writer, news_items, search_phrase)
    writer.save_excel(file_path)


if __name__ == '__main__':
    main()
